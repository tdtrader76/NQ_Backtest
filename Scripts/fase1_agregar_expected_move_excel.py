"""
Script para agregar Expected Move al Excel - FASE 1.3
Versión: 1.0
Fecha: 2025-01-26

Descripción:
    Agrega columnas EMH, EML y ExpRange al archivo Excel existente
    Datos_Diarios_por_Año.xlsx usando la fórmula del indicador RyFHEM.cs

Fórmula Expected Move:
    - EMH = Open + (avgBullishMove * 0.682)
    - EML = Open - (avgBearishMove * 0.682)
    - ExpRange = EMH - EML
    
    Donde:
    - avgBullishMove = promedio de rangos (High-Low) de días alcistas (Close > Open)
    - avgBearishMove = promedio de rangos (High-Low) de días bajistas (Close <= Open)
    - 0.682 = multiplicador de desviación estándar (68.2%)
"""

import pandas as pd
import numpy as np
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('../Logs/fase1_agregar_expected_move.log'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

EXCEL_PATH = Path("..") / "Resultados" / "Fase1" / "Datos_Diarios_por_Año.xlsx"
DATA_PATH = Path("..") / "Procesados" / "NQ_Daily_2020-2025_Limpio.csv"
RANGE_MULTIPLIER = 0.682
DEFAULT_LOOKBACK = 21

def cargar_datos():
    logger.info("="*80)
    logger.info("CARGANDO DATOS")
    logger.info("="*80)
    
    try:
        df = pd.read_csv(DATA_PATH)
        df['Date'] = pd.to_datetime(df['Date'])
        
        logger.info(f"OK Datos cargados: {len(df)} registros")
        logger.info(f"Período: {df['Date'].min()} a {df['Date'].max()}")
        
        return df
    
    except Exception as e:
        logger.error(f"ERROR al cargar datos: {str(e)}")
        raise

def calcular_expected_move(df, lookback=DEFAULT_LOOKBACK):
    logger.info("\n" + "="*80)
    logger.info(f"CALCULANDO EXPECTED MOVE (Lookback={lookback})")
    logger.info("="*80)
    
    df = df.copy()
    
    df['IsBullish'] = df['Close'] > df['Open']
    df['Range'] = df['High'] - df['Low']
    
    df['BullishAvg'] = 0.0
    df['BearishAvg'] = 0.0
    df['EMH'] = 0.0
    df['EML'] = 0.0
    df['ExpRange'] = 0.0
    
    for i in range(len(df)):
        if i < lookback:
            continue
        
        start_idx = i - lookback
        end_idx = i
        
        historical_window = df.iloc[start_idx:end_idx]
        
        bullish_days = historical_window[historical_window['IsBullish']]
        bearish_days = historical_window[~historical_window['IsBullish']]
        
        if len(bullish_days) > 0:
            df.loc[df.index[i], 'BullishAvg'] = bullish_days['Range'].mean()
        
        if len(bearish_days) > 0:
            df.loc[df.index[i], 'BearishAvg'] = bearish_days['Range'].mean()
        
        emh_raw = df.loc[df.index[i], 'Open'] + (df.loc[df.index[i], 'BullishAvg'] * RANGE_MULTIPLIER)
        eml_raw = df.loc[df.index[i], 'Open'] - (df.loc[df.index[i], 'BearishAvg'] * RANGE_MULTIPLIER)
        
        df.loc[df.index[i], 'EMH'] = round_to_nearest_quarter(emh_raw)
        df.loc[df.index[i], 'EML'] = round_to_nearest_quarter(eml_raw)
        df.loc[df.index[i], 'ExpRange'] = df.loc[df.index[i], 'EMH'] - df.loc[df.index[i], 'EML']
    
    valid_em = df[df['EMH'] > 0]
    logger.info(f"\nEstadísticas Expected Move:")
    logger.info(f"  Días con EM válido: {len(valid_em)} de {len(df)}")
    logger.info(f"  EMH promedio: {valid_em['EMH'].mean():.2f}")
    logger.info(f"  EML promedio: {valid_em['EML'].mean():.2f}")
    logger.info(f"  ExpRange promedio: {valid_em['ExpRange'].mean():.2f}")
    
    return df

def round_to_nearest_quarter(price):
    if price <= 0:
        return price
    
    quarters = np.ceil(price * 4)
    return quarters / 4

def actualizar_excel(df):
    logger.info("\n" + "="*80)
    logger.info("ACTUALIZANDO ARCHIVO EXCEL")
    logger.info("="*80)
    
    try:
        wb = load_workbook(EXCEL_PATH)
        
        df['Year'] = df['Date'].dt.year
        
        for year in sorted(df['Year'].unique()):
            if str(year) not in wb.sheetnames:
                logger.warning(f"WARN Hoja {year} no encontrada, saltando...")
                continue
            
            ws = wb[str(year)]
            df_year = df[df['Year'] == year].copy()
            
            header_row = 1
            headers = [cell.value for cell in ws[header_row]]
            
            if 'EMH' not in headers:
                emh_col = len(headers) + 1
                ws.cell(header_row, emh_col, 'EMH')
                ws.cell(header_row, emh_col).font = Font(bold=True)
                ws.cell(header_row, emh_col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:
                emh_col = headers.index('EMH') + 1
            
            if 'EML' not in headers:
                eml_col = len(headers) + 2 if 'EMH' not in headers else emh_col + 1
                ws.cell(header_row, eml_col, 'EML')
                ws.cell(header_row, eml_col).font = Font(bold=True)
                ws.cell(header_row, eml_col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:
                eml_col = headers.index('EML') + 1
            
            if 'ExpRange' not in headers:
                exprange_col = eml_col + 1
                ws.cell(header_row, exprange_col, 'ExpRange')
                ws.cell(header_row, exprange_col).font = Font(bold=True)
                ws.cell(header_row, exprange_col).fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            else:
                exprange_col = headers.index('ExpRange') + 1
            
            for idx, row_data in df_year.iterrows():
                row_num = list(df_year.index).index(idx) + 2
                
                emh_value = row_data['EMH'] if row_data['EMH'] > 0 else ''
                eml_value = row_data['EML'] if row_data['EML'] > 0 else ''
                exprange_value = row_data['ExpRange'] if row_data['ExpRange'] > 0 else ''
                
                ws.cell(row_num, emh_col, emh_value)
                ws.cell(row_num, eml_col, eml_value)
                ws.cell(row_num, exprange_col, exprange_value)
                
                if emh_value != '':
                    ws.cell(row_num, emh_col).number_format = '#,##0.00'
                    ws.cell(row_num, eml_col).number_format = '#,##0.00'
                    ws.cell(row_num, exprange_col).number_format = '#,##0.00'
            
            logger.info(f"OK Hoja {year} actualizada: {len(df_year)} registros")
        
        wb.save(EXCEL_PATH)
        logger.info(f"\n✅ Excel actualizado exitosamente: {EXCEL_PATH}")
        
    except Exception as e:
        logger.error(f"ERROR al actualizar Excel: {str(e)}")
        raise

def main():
    logger.info("\n" + "="*80)
    logger.info("INICIO - AGREGAR EXPECTED MOVE A EXCEL")
    logger.info("="*80)
    logger.info(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    try:
        df = cargar_datos()
        
        df = calcular_expected_move(df)
        
        actualizar_excel(df)
        
        logger.info("\n" + "="*80)
        logger.info("✅ PROCESO COMPLETADO EXITOSAMENTE")
        logger.info("="*80)
        
    except Exception as e:
        logger.error(f"\n❌ ERROR EN EL PROCESO: {str(e)}")
        raise

if __name__ == "__main__":
    main()
