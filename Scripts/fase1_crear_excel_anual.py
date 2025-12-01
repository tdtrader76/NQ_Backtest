"""
Script de Organización en Excel por Años - FASE 1 Tarea 1.2
Versión: 1.0
Fecha: 2025-01-26
Autor: Sistema Backtesting NASDAQ

Descripción:
    Crea archivo Excel con hojas separadas por año para validación visual y manual.
    Incluye formato condicional para resaltar posibles errores.

Según PLAN_BACKTESTING_NASDAQ.md:
    - Hoja por cada año (2020-2025)
    - Hoja RESUMEN con estadísticas generales
    - Columnas de validación
    - Formato condicional (errores en rojo)
"""

import pandas as pd
from pathlib import Path
import logging
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('../Logs/fase1_excel_anual.log'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)

# Constantes
DATA_PATH = Path("..") / "Procesados" / "NQ_Daily_2020-2025_Limpio.csv"
OUTPUT_PATH = Path("..") / "Resultados" / "Fase1"
OUTPUT_FILE = OUTPUT_PATH / "Datos_Diarios_por_Año.xlsx"

def cargar_datos():
    """
    Carga los datos diarios limpios

    Returns:
        DataFrame con datos diarios
    """
    logger.info("="*80)
    logger.info("CARGANDO DATOS")
    logger.info("="*80)

    try:
        df = pd.read_csv(DATA_PATH)
        df['Date'] = pd.to_datetime(df['Date'])
        df['Year'] = df['Date'].dt.year

        logger.info(f"OK Datos cargados: {len(df)} registros")
        logger.info(f"Años presentes: {sorted(df['Year'].unique())}")

        return df

    except Exception as e:
        logger.error(f"ERROR al cargar datos: {str(e)}")
        raise

def agregar_columnas_validacion(df):
    """
    Agrega columnas de validación al DataFrame

    Args:
        df: DataFrame con datos diarios

    Returns:
        DataFrame con columnas de validación adicionales
    """
    logger.info("\n" + "="*80)
    logger.info("AGREGANDO COLUMNAS DE VALIDACION")
    logger.info("="*80)

    # Validación básica High >= Low
    df['H>=L'] = df['High'] >= df['Low']

    # Validación de rango (Open y Close dentro de High-Low)
    df['Valid_Range'] = (
        (df['Open'].between(df['Low'], df['High'])) &
        (df['Close'].between(df['Low'], df['High']))
    )

    # Calcular rango diario
    df['Range'] = df['High'] - df['Low']

    # Calcular return diario
    df['Return_%'] = df['Close'].pct_change() * 100

    logger.info("OK Columnas de validacion agregadas")
    logger.info(f"  - H>=L: {df['H>=L'].sum()} validos de {len(df)}")
    logger.info(f"  - Valid_Range: {df['Valid_Range'].sum()} validos de {len(df)}")

    return df

def crear_hoja_resumen(df):
    """
    Crea DataFrame con resumen estadístico por año

    Args:
        df: DataFrame con datos diarios

    Returns:
        DataFrame con resumen por año
    """
    logger.info("\n" + "="*80)
    logger.info("CREANDO HOJA RESUMEN")
    logger.info("="*80)

    # Agrupar por año y calcular estadísticas
    resumen = df.groupby('Year').agg({
        'Date': 'count',
        'Close': ['mean', 'std', 'min', 'max'],
        'Volume': ['sum', 'mean'],
        'Range': 'mean',
        'Return_%': ['mean', 'std'],
        'Valid_Range': 'sum'
    }).round(2)

    # Renombrar columnas para mayor claridad
    resumen.columns = [
        'Dias_Trading',
        'Close_Mean', 'Close_Std', 'Close_Min', 'Close_Max',
        'Volume_Total', 'Volume_Mean',
        'Range_Mean',
        'Return_Mean_%', 'Return_Std_%',
        'Registros_Validos'
    ]

    # Calcular porcentaje de validez
    resumen['Validez_%'] = (resumen['Registros_Validos'] / resumen['Dias_Trading'] * 100).round(2)

    logger.info("OK Hoja resumen creada")
    logger.info(f"\n{resumen}")

    return resumen

def crear_excel_con_formato(df, resumen):
    """
    Crea archivo Excel con hojas por año y formato condicional

    Args:
        df: DataFrame completo con validaciones
        resumen: DataFrame con resumen por año
    """
    logger.info("\n" + "="*80)
    logger.info("CREANDO ARCHIVO EXCEL")
    logger.info("="*80)

    # Crear carpeta si no existe
    OUTPUT_PATH.mkdir(parents=True, exist_ok=True)

    # Crear Excel con pandas
    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl') as writer:
        # Hoja 1: RESUMEN
        resumen.to_excel(writer, sheet_name='RESUMEN')
        logger.info("OK Hoja RESUMEN creada")

        # Hojas por año
        for year in sorted(df['Year'].unique()):
            df_year = df[df['Year'] == year].copy()

            # Seleccionar columnas para mostrar
            columnas_display = [
                'Date', 'Open', 'High', 'Low', 'Close', 'Volume',
                'Range', 'Return_%', 'H>=L', 'Valid_Range'
            ]
            df_year = df_year[columnas_display]

            # Formatear fecha como string para mejor visualización
            df_year['Date'] = df_year['Date'].dt.strftime('%Y-%m-%d')

            # Exportar a Excel
            df_year.to_excel(writer, sheet_name=str(year), index=False)
            logger.info(f"OK Hoja {year} creada ({len(df_year)} registros)")

    # Aplicar formato condicional
    aplicar_formato_excel(df, resumen)

    logger.info(f"\nOK Excel guardado en: {OUTPUT_FILE}")

def aplicar_formato_excel(df, resumen):
    """
    Aplica formato condicional al archivo Excel

    Args:
        df: DataFrame completo
        resumen: DataFrame resumen
    """
    logger.info("\n" + "="*80)
    logger.info("APLICANDO FORMATO CONDICIONAL")
    logger.info("="*80)

    # Cargar el workbook
    wb = load_workbook(OUTPUT_FILE)

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    error_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Formatear hoja RESUMEN
    ws_resumen = wb['RESUMEN']
    logger.info("Formateando hoja RESUMEN...")

    # Headers
    for cell in ws_resumen[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Ajustar anchos de columna
    for col in ws_resumen.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 20)
        ws_resumen.column_dimensions[column].width = adjusted_width

    # Formatear hojas por año
    for year in sorted(df['Year'].unique()):
        ws = wb[str(year)]
        df_year = df[df['Year'] == year]

        logger.info(f"Formateando hoja {year}...")

        # Headers
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border

        # Formatear datos
        for row_idx in range(2, len(df_year) + 2):
            row_data = df_year.iloc[row_idx - 2]

            # Aplicar bordes a todas las celdas
            for col_idx in range(1, 11):  # 10 columnas
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                cell.alignment = Alignment(horizontal="right", vertical="center")

            # Resaltar errores en rojo
            if not row_data['Valid_Range']:
                for col_idx in range(1, 11):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = error_fill

            # Formatear números
            ws.cell(row=row_idx, column=2).number_format = '#,##0.00'  # Open
            ws.cell(row=row_idx, column=3).number_format = '#,##0.00'  # High
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00'  # Low
            ws.cell(row=row_idx, column=5).number_format = '#,##0.00'  # Close
            ws.cell(row=row_idx, column=6).number_format = '#,##0'     # Volume
            ws.cell(row=row_idx, column=7).number_format = '#,##0.00'  # Range
            ws.cell(row=row_idx, column=8).number_format = '0.00%'     # Return_%

        # Ajustar anchos de columna
        column_widths = {
            'A': 12,  # Date
            'B': 12,  # Open
            'C': 12,  # High
            'D': 12,  # Low
            'E': 12,  # Close
            'F': 12,  # Volume
            'G': 12,  # Range
            'H': 12,  # Return_%
            'I': 8,   # H>=L
            'J': 12   # Valid_Range
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # Congelar primera fila
        ws.freeze_panes = 'A2'

    # Guardar cambios
    wb.save(OUTPUT_FILE)
    logger.info("OK Formato aplicado correctamente")

def generar_resumen_final(df, resumen):
    """
    Genera resumen final de la tarea

    Args:
        df: DataFrame completo
        resumen: DataFrame resumen
    """
    logger.info("\n" + "="*80)
    logger.info("RESUMEN FINAL")
    logger.info("="*80)

    logger.info(f"\nArchivo creado: {OUTPUT_FILE}")
    logger.info(f"Tamaño: {OUTPUT_FILE.stat().st_size / 1024:.2f} KB")

    logger.info("\n--- HOJAS CREADAS ---")
    logger.info("  1. RESUMEN - Estadisticas por año")

    for idx, year in enumerate(sorted(df['Year'].unique()), start=2):
        count = len(df[df['Year'] == year])
        logger.info(f"  {idx}. {year} - {count} registros")

    logger.info("\n--- VALIDACION MANUAL ---")
    logger.info("Pasos recomendados:")
    logger.info("  [ ] Revisar registros resaltados en rojo")
    logger.info("  [ ] Verificar primeros y ultimos registros de cada año")
    logger.info("  [ ] Comprobar que no hay gaps inesperados")
    logger.info("  [ ] Validar que totales de RESUMEN son coherentes")

def main():
    """
    Función principal
    """
    try:
        logger.info("\n" + "="*80)
        logger.info("FASE 1 - TAREA 1.2: ORGANIZACION EN EXCEL POR AÑOS")
        logger.info("="*80)

        # 1. Cargar datos
        df = cargar_datos()

        # 2. Agregar columnas de validación
        df = agregar_columnas_validacion(df)

        # 3. Crear hoja resumen
        resumen = crear_hoja_resumen(df)

        # 4. Crear Excel con formato
        crear_excel_con_formato(df, resumen)

        # 5. Resumen final
        generar_resumen_final(df, resumen)

        logger.info("\n" + "="*80)
        logger.info("TAREA 1.2 COMPLETADA EXITOSAMENTE")
        logger.info("="*80)

    except Exception as e:
        logger.error(f"ERROR en tarea 1.2: {str(e)}")
        raise

if __name__ == "__main__":
    main()
